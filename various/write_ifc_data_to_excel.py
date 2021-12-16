import os
import bpy
import openpyxl
import ifcopenshell
import xlsxwriter

import logging
import blenderbim.bim.import_ifc

from openpyxl import load_workbook
from blenderbim.bim.ifc import IfcStore
import blenderbim.tool as tool

  
def write_to_excel_from_ifc(ifc_file,excel_file):
    
    ifc_file = ifcopenshell.open(ifc_file)
    products = ifc_file.by_type('IfcProduct')
    
    workbook_xlsx = xlsxwriter.Workbook(excel_file)
    worksheet_xlsx = workbook_xlsx.add_worksheet('IfcProduct')
    #worksheet_xlsx.autofilter('A1:J1' + str(len(products)) )
    worksheet_xlsx.add_table('A1:J' + str(len(products)))
    
    product_entity_list = [ ['A1','GlobalId'],
                            ['B1','IfcProduct'],
                            ['C1','IfcBuildingStorey'],
                            ['D1','Name'],
                            ['E1','Type'],
                            ['F1','Classification'],
                            ['G1','IfcMaterial'],
                            ['H1','IsExternal'],
                            ['I1','LoadBearing'],
                            ['J1','FireRating']]
    
    for i, product_entity in enumerate(product_entity_list):
        worksheet_xlsx.write(product_entity[0], product_entity[1])
        worksheet_xlsx.set_column(i, 1, 30)
    
    for i, product in enumerate(products):
        worksheet_xlsx.write('A' + str(i+2), str(product.GlobalId))
        worksheet_xlsx.write('B' + str(i+2),(get_ifcproduct(ifcproduct=product)[0]))
        worksheet_xlsx.write('C' + str(i+2),(get_ifcbuildingstorey(ifcproduct=product)[0]))
        worksheet_xlsx.write('D' + str(i+2), str(product.Name))
        worksheet_xlsx.write('E' + str(i+2),(get_ifcproducttype_name(ifcproduct=product)[0]))
        worksheet_xlsx.write('F' + str(i+2),(get_classification_code(ifcproduct=product)[0]))
        worksheet_xlsx.write('G' + str(i+2),(get_materials(ifcproduct=product)[0]))
        worksheet_xlsx.write('H' + str(i+2),(get_isexternal(ifcproduct=product)[0][0]))
        worksheet_xlsx.write('I' + str(i+2),(get_loadbearing(ifcproduct=product)[0][0]))
        worksheet_xlsx.write('J' + str(i+2),(get_firerating(ifcproduct=product)[0][0]))
            
    workbook_xlsx.close()
    os.startfile(excel_file)
        
    
def get_ifcproduct(ifcproduct):
    product_list = []
    
    if ifcproduct:
        product_list.append(ifcproduct.is_a())

    return product_list 

def get_ifcproducttype_name(ifcproduct):
    type_name_list = []
    
    if ifcproduct:
        type_name_list.append(ifcproduct.ObjectType)
        
    if len(type_name_list) == 0:
        type_name_list.append('N/A')
        
    if type_name_list[0] == None:
        type_name_list.append('N/A')
        
    return type_name_list
   
def get_ifcbuildingstorey(ifcproduct):
    building_storey_list = []
     
    try:
        if ifcproduct:
            if ifcproduct.ContainedInStructure:            
                if ifcproduct.ContainedInStructure[0].RelatingStructure.is_a() == 'IfcBuildingStorey':
                    building_storey_list.append(ifcproduct.ContainedInStructure[0].RelatingStructure.Name)
    except:
        pass
    else:
        pass

    #IfcOpeningElement should not be linked directly to the spatial structure of the project,
    #i.e. the inverse relationship ContainedInStructure shall be NIL.
    #It is assigned to the spatial structure through the elements it penetrates.
    if len(building_storey_list) == 0:
        building_storey_list.append('N/A')
         
    return building_storey_list 

def get_classification_code(ifcproduct):
    
    #Classifications of an object may be referenced from an external source rather than being
    #contained within the IFC model. This is done through the IfcClassificationReference class.
    
    assembly_code_list = []

    if ifcproduct.HasAssociations:
        if ifcproduct.HasAssociations[0].is_a() == 'IfcRelAssociatesClassification':
            assembly_code_list.append(ifcproduct.HasAssociations[0].RelatingClassification.ItemReference)
            
        if ifcproduct.HasAssociations[0].is_a() == 'IfcRelAssociatesMaterial':
            for i in ifcproduct.HasAssociations:
                if i.is_a() == 'IfcRelAssociatesClassification' :
                    assembly_code_list.append(i.RelatingClassification.ItemReference)
            
                                           
    if len(assembly_code_list) == 0:
        assembly_code_list.append('N/A')
        
    if assembly_code_list[0] == None:
        assembly_code_list.append('N/A')
        
    return assembly_code_list

def get_materials(ifcproduct):
    material_list = []
    
    if ifcproduct.HasAssociations:
        for i in ifcproduct.HasAssociations:
            if i.is_a('IfcRelAssociatesMaterial'):
                
                if i.RelatingMaterial.is_a('IfcMaterial'):
                    material_list.append(i.RelatingMaterial.Name)
                   
                if i.RelatingMaterial.is_a('IfcMaterialList'):
                    for materials in i.RelatingMaterial.Materials:
                        material_list.append(materials.Name)
                         
                if i.RelatingMaterial.is_a('IfcMaterialLayerSetUsage'):
                    for materials in i.RelatingMaterial.ForLayerSet.MaterialLayers:
                        material_list.append(materials.Material.Name)
                        
                else:
                    pass
                      
    if len(material_list) == 0:
        material_list.append('N/A')
       
    joined_material_list = ' | '.join(material_list)
                         
    return [joined_material_list] 

def get_isexternal(ifcproduct):
    externality_list = []
 
    if ifcproduct.IsDefinedBy:
        if ifcproduct.IsDefinedBy[0]:
            if ifcproduct.IsDefinedBy[0].is_a() == 'IfcRelDefinesByProperties':
                if ifcproduct.IsDefinedBy[0].RelatingPropertyDefinition.is_a() == 'IfcPropertySet':
                    for ifcproperty in ifcproduct.IsDefinedBy[0].RelatingPropertyDefinition.HasProperties:
                        if ifcproperty.Name == 'IsExternal':
                            #externality_list.append(ifcproperty.Name)
                            externality_list.append(ifcproperty.NominalValue[0])
                        else:
                            pass
                                    
    if len(externality_list) == 0:
        externality_list.append('N/A')  

    return [externality_list]

def get_loadbearing(ifcproduct):
    load_bearing_list = []
    
    if ifcproduct.IsDefinedBy:
        if ifcproduct.IsDefinedBy[0]:
            if ifcproduct.IsDefinedBy[0].is_a() == 'IfcRelDefinesByProperties':
                if ifcproduct.IsDefinedBy[0].RelatingPropertyDefinition.is_a() == 'IfcPropertySet':
                    for ifcproperty in ifcproduct.IsDefinedBy[0].RelatingPropertyDefinition.HasProperties:
                        if ifcproperty.Name == 'LoadBearing':
                            load_bearing_list.append(ifcproperty.NominalValue[0])
                        else:
                            pass
                                    
    if len(load_bearing_list) == 0:
        load_bearing_list.append('N/A')
                         
    return [load_bearing_list]

def get_firerating(ifcproduct):
    fire_rating_list = []    
    
    if ifcproduct.IsDefinedBy:
        if ifcproduct.IsDefinedBy[0]:
            if ifcproduct.IsDefinedBy[0].is_a() == 'IfcRelDefinesByProperties':
                if ifcproduct.IsDefinedBy[0].RelatingPropertyDefinition.is_a() == 'IfcPropertySet':
                    for ifcproperty in ifcproduct.IsDefinedBy[0].RelatingPropertyDefinition.HasProperties:
                        if ifcproperty.Name == 'FireRating':
                            fire_rating_list.append(ifcproperty.Name)
                            fire_rating_list.append(ifcproperty.NominalValue[0])
                        else:
                            pass
                                    
    if len(fire_rating_list) == 0:
        fire_rating_list.append('N/A')
                         
    return [fire_rating_list]


def get_filtered_data_from_excel(excel_file):
    workbook_openpyxl = load_workbook(excel_file)
    worksheet_openpyxl = workbook_openpyxl['IfcProduct'] 
    
    global_id_filtered_list = []

    for row in worksheet_openpyxl:     
        if worksheet_openpyxl.row_dimensions[row[0].row].hidden == False:
            for cell in row:  
                if cell in worksheet_openpyxl['A']:  
                    global_id_filtered_list.append(cell.value)
                    
    return global_id_filtered_list[1:]
                
                                            
def select_IFC_elements_in_blender(guid_list, excel_file):
    
    if excel_file is not None:
        os.startfile(excel_file)
    
    outliner = next(a for a in bpy.context.screen.areas if a.type == "OUTLINER") 
    outliner.spaces[0].show_restrict_column_viewport = not outliner.spaces[0].show_restrict_column_viewport
    
    bpy.ops.object.select_all(action='DESELECT')
  
    for obj in bpy.context.view_layer.objects:
        element = tool.Ifc.get_entity(obj)
        if element is None:        
            obj.hide_viewport = True
            continue
        data = element.get_info()
      
        obj.hide_viewport = data.get("GlobalId", False) not in guid_list

    bpy.ops.object.select_all(action='SELECT')
  
def unhide_all():
    
    for obj in bpy.data.objects:
        obj.hide_viewport = False        
   
       
excel_file_path = (os.path.dirname(IfcStore.path) + '\\' + (os.path.basename(IfcStore.path).replace('.ifc','.xlsx')) )

#1 export the excel first
write_to_excel_from_ifc(ifc_file=IfcStore.path, excel_file=excel_file_path)

#2 check if excel is running and saved before using this function
#select_IFC_elements_in_blender(guid_list=get_filtered_data_from_excel(excel_file=excel_file_path), excel_file=excel_file_path )   

#reset hide isolate
#unhide_all()

#mini backlog
#0. remove spatial elements to another column ifcbuildingstorey
#1. export propertyset to a sheet
#2. export materials to a sheet
#3. export quantities to a sheet
#4. find out how to bundle python dependencies in a blender add-on
#5. small user interface with pyqt or blender
#6. check if excel is installed on the system or running
#7. refactor function write_to_excel()
#8. write back data from excel into ifc